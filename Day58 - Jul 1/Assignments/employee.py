import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
import csv
import xlsxwriter
from fpdf import FPDF
import os
import openpyxl

class Employee:
    """
    A class to represent an employee.
    """

    def __init__(self, name, dob, phone, email):
        """
        Initialize the employee with name, date of birth, phone, and email.

        :param name: Name of the employee
        :param dob: Date of birth of the employee in YYYY-MM-DD format
        :param phone: Phone number of the employee
        :param email: E-Mail of the employee
        """
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email
        self.age = self.calculate_age()

    def calculate_age(self):
        """
        Calculate the age of the employee based on the date of birth.

        :return: Age in years
        """
        birth_date = datetime.strptime(self.dob, "%Y-%m-%d")
        today = datetime.today()
        age = relativedelta(today, birth_date).years
        return age


def validate_name(name):
    """
    Validate the employee's name.

    :param name: Name of the employee
    :return: True if valid, raises ValueError otherwise
    """
    if not name.isalpha():
        raise ValueError("Name should contain only alphabets.")
    return True

def validate_dob(dob):
    """
    Validate the date of birth.

    :param dob: Date of birth in YYYY-MM-DD format
    :return: True if valid, raises ValueError otherwise
    """
    try:
        datetime.strptime(dob, "%Y-%m-%d")
        return True
    except ValueError:
        raise ValueError("Date of birth should be in YYYY-MM-DD format.")

def validate_phone(phone):
    """
    Validate the phone number.

    :param phone: Phone number of the employee
    :return: True if valid, raises ValueError otherwise
    """
    if not re.match(r'^\d{10}$', phone):
        raise ValueError("Phone number should be a 10-digit number.")
    return True

def validate_email(email):
    """
    Validate the email address.

    :param email: E-Mail of the employee
    :return: True if valid, raises ValueError otherwise
    """
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        raise ValueError("Invalid email address format.")
    return True

def get_employee_details():
    """
    Get employee details from user input.

    :return: A tuple containing the employee's name, date of birth, phone, and email
    """
    while True:
        try:
            name = input("Enter the employee's name: ")
            validate_name(name)
            dob = input("Enter the employee's date of birth (YYYY-MM-DD): ")
            validate_dob(dob)
            phone = input("Enter the employee's phone number: ")
            validate_phone(phone)
            email = input("Enter the employee's email: ")
            validate_email(email)
            return name, dob, phone, email
        except ValueError as ve:
            print(f"Error: {ve}")
            print("Please try again.")

def save_to_text(employees):
    """
    Save employee details to a text file.

    :param employees: List of Employee objects
    """
    with open('Assignments/employees.txt', 'w') as file:
        for emp in employees:
            file.write(f"Name: {emp.name}\nDOB: {emp.dob}\nPhone: {emp.phone}\nEmail: {emp.email}\nAge: {emp.age}\n\n")
    print("Employee details saved to employees.txt")

def save_to_excel(employees):
    """
    Save employee details to an Excel file.

    :param employees: List of Employee objects
    """
    workbook = xlsxwriter.Workbook('Assignments/employees.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', 'Name')
    worksheet.write('B1', 'DOB')
    worksheet.write('C1', 'Phone')
    worksheet.write('D1', 'Email')
    worksheet.write('E1', 'Age')

    row = 1
    for emp in employees:
        worksheet.write(row, 0, emp.name)
        worksheet.write(row, 1, emp.dob)
        worksheet.write(row, 2, emp.phone)
        worksheet.write(row, 3, emp.email)
        worksheet.write(row, 4, emp.age)
        row += 1

    workbook.close()
    print("Employee details saved to employees.xlsx")

def save_to_pdf(employees):
    """
    Save employee details to a PDF file.

    :param employees: List of Employee objects
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for emp in employees:
        pdf.cell(200, 10, txt=f"Name: {emp.name}", ln=True)
        pdf.cell(200, 10, txt=f"DOB: {emp.dob}", ln=True)
        pdf.cell(200, 10, txt=f"Phone: {emp.phone}", ln=True)
        pdf.cell(200, 10, txt=f"Email: {emp.email}", ln=True)
        pdf.cell(200, 10, txt=f"Age: {emp.age}", ln=True)
        pdf.cell(200, 10, txt="", ln=True)

    pdf.output("D:/PresidioTraining/Day58 - Jul 1/Assignments/employees.pdf")
    print("Employee details saved to employees.pdf")

def read_from_excel(file_path):
    """
    Read employee details from an Excel file and return a list of Employee objects.

    :param file_path: Path to the Excel file
    :return: List of Employee objects
    """
    employees = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, values_only=True):
        # if len(row) != 4:
        #     print(f"Skipping invalid data row: {row} - Incorrect number of values.")
        #     continue

        name, dob, phone, email, age = row
        try:
            validate_name(name)
            validate_dob(dob)
            validate_phone(phone)
            validate_email(email)
            employee = Employee(name, dob, phone, email)
            employees.append(employee)
        except ValueError as ve:
            print(f"Skipping invalid data row: {ve}")

    return employees

def printEmployees(employees):
    for employee in employees:
        print(f"\nname: {employee.name}")
        print(f"dob: {employee.dob}")
        print(f"phone: {employee.phone}")
        print(f"email: {employee.email}\n")

def main():
    """
    Main function to run the application.
    """
    employees = []

    while True:
        print("\nEmployee Management System")
        print("1. Add Employee")
        print("2. Bulk Read from Excel")
        print("3. Save Employee Details")
        print("4. Exit")

        choice = input("Choose an option: ")

        if choice == '1':
            name, dob, phone, email = get_employee_details()
            employee = Employee(name, dob, phone, email)
            employees.append(employee)
        elif choice == '2':
            file_path = input("Enter the path to the Excel file: ")
            if os.path.exists(file_path):
                new_employees = read_from_excel(file_path)
                employees.extend(new_employees)
                print(f"{len(new_employees)} employees added from the Excel file.")
                printEmployees(employees)
            else:
                print("Invalid file path. Please try again.")
        elif choice == '3':
            if not employees:
                print("No employee details to save.")
                continue

            print("Save Options:")
            print("1. Text File")
            print("2. Excel File")
            print("3. PDF File")

            save_choice = input("Choose an option to save: ")

            if save_choice == '1':
                save_to_text(employees)
            elif save_choice == '2':
                save_to_excel(employees)
            elif save_choice == '3':
                save_to_pdf(employees)
            else:
                print("Invalid option. Please try again.")
        elif choice == '4':
            print("Exiting the application. Goodbye!")
            break
        else:
            print("Invalid option. Please try again.")

if __name__ == "__main__":
    main()